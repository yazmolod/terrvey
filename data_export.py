import openpyxl as xl
from string import ascii_uppercase as ABC


def to_excel(connection, structure, path):
    wb = xl.Workbook()
    wb.active.title = 'Данные'
    cursor = connection.cursor()
    cursor.execute('SELECT * FROM "view_Respondents"')
    data = cursor.fetchall()
    for r in data:
        wb['Данные'].append(r)
    for q in structure:
        table_name = q['table_name']
        quest = q['quest']
        qtype = q['type']
        if 'Quest' in table_name:
            cursor.execute('SELECT * FROM "view_{}"'.format(table_name))
            data = cursor.fetchall()
            if 'matrix' in qtype:
                print(data)
            if 'text' in qtype:
                wb.create_sheet(table_name)
                wb[table_name]['A1'].value = quest
                for d in data:
                    wb[table_name].append(d)
            else:
                ws = wb['Данные']
                ws.cell(ws.max_row + 2, 1, quest)
                for d in data:
                    row = ws.max_row + 1
                    if 'matrix' in qtype:
                        ws.cell(row, 1, d[0])
                        d = d[1:]
                        for i in range(len(d)):
                            ws.cell(row, i * 2 + 2, d[i])
                            ws.cell(row, i * 2 + 3,
                                    '=={}{}/$B$1*100'.format(ABC[i * 2 + 1], row)).number_format = '0.0'

                    else:
                        ws.append(d)
                        srow = str(row)
                        ws['C'+srow].value = '==B'+srow+'/$B$1*100'
                        ws['C'+srow].number_format = '0.0'
    wb.save(path)

def to_images(connection, structure, path):
	print(path)
