import sys
import json
import os

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi

import openpyxl

from configBuilder import ConfigConstructor

from Grand_kotlovan import algoritm


class workTableModel(QAbstractTableModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.columns = ("N", "Name", "Unit", "Value")
        self.innerData = []

    def addData(self, data):
        self.innerData += data
        self.layoutChanged.emit()

    def deleteRow(self, row):
        del self.innerData[row]
        self.layoutChanged.emit()

    def rowCount(self, *args):
        return len(self.innerData)

    def columnCount(self, *args):
        return len(self.columns)

    def data(self, index, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            i = index.row()
            j = index.column()
            column_name = self.columns[j]
            return self.innerData[i].get(column_name, "")
        else:
            return QVariant()


class MainApp(QMainWindow):
    structFilePath = "./config.json"

    def __init__(self):
        """Инициализация главного окна"""
        super().__init__()
        self.loadUI()
        self.loadConfig()
        self.initMenu()
        self.initCentralWidget()

    def loadUI(self):
        self.resize(725, 575)
        self.setSizePolicy = QSizePolicy.Fixed

    def loadConfig(self):
        """Загружаем файл конфигурации с прописанными работами"""
        if os.path.exists(self.structFilePath):
            with open(self.structFilePath, 'r', encoding='utf-8') as file:
                self.Config = json.load(file)
        else:
            QMessageBox.warning(
                self, "Ошибка", "Отсутствует файл конфигурации")
            self.Config = None

    def initMenu(self):
        editConfigAction = QAction("&Редактировать", self)
        editConfigAction.triggered.connect(self.editConfig)

        fileMenu = self.menuBar().addMenu("&Конфигурация")
        fileMenu.addAction(editConfigAction)

    def initCentralWidget(self):
        self.setCentralWidget(CentralWidget(self))

    def editConfig(self):
        dialog = ConfigConstructor(self.Config)
        dialog.exec()


class CentralWidget(QWidget):

    def __init__(self, parent=None):
        """Инициализация центрального виджета"""
        super().__init__(parent=parent)
        self.tableModel = workTableModel()
        self.loadUI()

    def loadUI(self):
        """Загрузка интерфейса"""
        loadUi(r'./ui/CentralWidget.ui', self)
        self.excelExportButton.clicked.connect(self.export_clicked)
        # Загрузка всех кнопок в соответствии с файлом конфига
        # и добавление их в коллекцию для доступа в будущем
        self.btnCollection = []
        for btnData in self.parent().Config:
            btnIconPath = btnData.get("IconPath", "")
            if os.path.exists(btnIconPath):
                icon = QIcon(btnIconPath)
            else:
                icon = QIcon()
            btnText = btnData.get("Name")

            btn = QPushButton(icon, btnText, self)
            btn.clicked.connect(self.button_clicked)
            self.workButtons.layout().addWidget(btn)
            self.btnCollection.append(btn)
        # обработка TableView с результами
        self.tableView.setModel(self.tableModel)
        self.tableView.contextMenuEvent = self.menuEvents

    def menuEvents(self, event):
        self.menu = QMenu(self)
        renameAction = QAction('Удалить', self)
        renameAction.triggered.connect(lambda: self.deleteModelItem(event))
        self.menu.addAction(renameAction)
        # add other required actions
        self.menu.popup(QCursor.pos())

    def deleteModelItem(self, event):
        row = self.tableView.rowAt(event.pos().y())
        self.tableModel.deleteRow(row)

    def button_clicked(self):
        btnText = self.sender().text()
        inputDataConfig = [i for i in self.parent(
        ).Config if i['Name'] == btnText][0]['Sections']
        dialog = InputDataDialog(inputDataConfig)
        ret = dialog.exec_()
        if ret == QDialog.Accepted:
            self.addNewWork(dialog.outputs)

    def addNewWork(self, tableResult):
        self.tableModel.addData(tableResult)
        self.tableView.update()

    def export_clicked(self):
        TEMPLATE_PATH = "./FormXLS.xlsm"
        RANGE_NAME = 'ТаблицаВедомость'
        data = self.tableModel.innerData
        fn = QFileDialog.getSaveFileName(parent=self, filter="(*.xlsm)")[0]
        if fn:
            wb = openpyxl.load_workbook(
                TEMPLATE_PATH, read_only=False, keep_vba=True)
            print(wb.sheetnames)
            ws = wb["Данные"]
            last_row = 0
            for i in range(len(data)):
                ws.cell(i + 2, 3).value = data[i]['Name']
                ws.cell(i + 2, 4).value = data[i]['Unit']
                ws.cell(i + 2, 5).value = data[i]['Value']
                last_row = i + 2
            rng = [dn for dn in wb.defined_names.definedName if dn.name == RANGE_NAME][0]
            rng.attr_text = '$'.join(rng.attr_text.split('$')[:-1] + [str(last_row)])
            wb.save(fn)
            QMessageBox.information(
                self, "Результат", "Файл успешно экспортирован")


class InputDataDialog(QDialog):

    def __init__(self, config):
        super().__init__()
        loadUi(r"./ui/inputDialog.ui", self)
        self.inputs = []
        self.outputs = []
        self.sectionData = None
        self.backButton.setEnabled(False)
        self.config = config
        self.loadSectionSheet(self.Stack.widget(
            0), [i.get("Name") for i in config])
        self.loadSignals()

    def loadSignals(self):
        self.nextButton.clicked.connect(self.nextButton_clicked)
        self.backButton.clicked.connect(self.backButton_clicked)
        self.acceptButton.clicked.connect(self.acceptButton_clicked)
        self.buttonGroup.buttonClicked.connect(self.radioButtonGroup_clicked)

    def loadSectionSheet(self, parent, data):
        layout = QVBoxLayout(parent)
        self.buttonGroup = QButtonGroup(parent)
        for section in data:
            radio = QRadioButton(section, parent)
            self.buttonGroup.addButton(radio)
            layout.addWidget(radio)

    def loadInputSheet(self, parent):
        data = self.sectionData["Inputs"]
        self.inputs.clear()
        layout = QGridLayout(parent)
        if self.sectionData["Name"] == "КЖ":
            vorButton = QPushButton("Запустить расчет котлована", self)
            vorButton.clicked.connect(self.vorStart)
            layout.addWidget(vorButton, 0, 0, 2, 0)
        for i in range(len(data)):
            nameLabel = QLabel(data[i].get("Name"), parent)
            measureLabel = QLabel(data[i].get("Unit"), parent)
            le = QDoubleSpinBox(parent)
            le.setDecimals(3)
            le.valueChanged.connect(self.updateOutputs)
            self.inputs.append(le)
            layout.addWidget(nameLabel, i+1, 0)
            layout.addWidget(le, i+1, 1)
            layout.addWidget(measureLabel, i+1, 2)

    def vorStart(self):
        vorWidget = algoritm.MyWin(self)
        vorWidget.exec_()

    def loadOutputSheet(self, parent):
        data = self.sectionData["Outputs"]
        layout = QGridLayout(parent)
        self.outputs.clear()
        for i in range(len(data)):
            name = data[i].get("Name")
            unit = data[i].get("Unit")
            nameLabel = QLabel(name, parent)
            measureLabel = QLabel(unit, parent)

            le = QDoubleSpinBox(parent)
            self.outputs.append({"Name": name, "Unit": unit, "UI": le})
            le.formula = data[i].get("Formula", "").replace(
                "[", "self.inputs[").replace("]", "].value()")
            le.setValue(0)

            le.setEnabled(False)
            layout.addWidget(nameLabel, i, 0)
            layout.addWidget(le, i, 1)
            layout.addWidget(measureLabel, i, 2)

    def updateOutputs(self, *args):
        for o in self.outputs:
            exec('global xxx; xxx = ' + o['UI'].formula)
            o['UI'].setValue(xxx)

    def radioButtonGroup_clicked(self, btn):
        inputWidget = self.Stack.widget(1)
        outputWidget = self.Stack.widget(2)
        inputLayout = inputWidget.layout()
        if inputLayout != None:
            for i in reversed(range(inputLayout.count())):
                inputLayout.itemAt(i).widget().setParent(None)
        self.sectionData = [
            i for i in self.config if i['Name'] == btn.text()][0]
        self.loadInputSheet(inputWidget)
        self.loadOutputSheet(outputWidget)

    def nextButton_clicked(self):
        curInd = self.Stack.currentIndex()
        self.Stack.setCurrentIndex(curInd + 1)
        self.backButton.setEnabled(True)
        self.nextButton.setEnabled(curInd != self.Stack.count() - 2)

    def backButton_clicked(self):
        curInd = self.Stack.currentIndex()
        self.Stack.setCurrentIndex(curInd - 1)
        self.nextButton.setEnabled(True)
        self.backButton.setEnabled(curInd - 1 != 0)

    def acceptButton_clicked(self):
        for i in self.outputs:
            i['Value'] = i['UI'].value()
        self.accept()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = MainApp()
    w.show()
    sys.exit(app.exec_())
